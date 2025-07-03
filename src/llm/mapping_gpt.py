import openai
import ast, sys
from os import path
from pathlib import Path
from typing import Union


PIG_PATH = Path.home() / "Desktop" / "pig_sal"

try:
    from synth import llm_pre
except:
    sys.path.append(path.dirname(path.dirname(path.abspath(__file__))))
    from synth import llm_pre

MODEL = ""
openai.api_key_path = ""


def ExtractLLM(libo: str, libn: str, answer: str) -> dict[str]:

    result = dict()

    # Extract function/method name
    llm_answer = answer.strip()
    result["codes"] = CodeExtract(llm_answer, libo, libn)

    return result

# current file: default setting (gpt3.5-turbo api)
def prepare(model: str, option: str = "default") -> dict[str, list[str]]:
    from openpyxl import load_workbook

    if option == "+slicing":
        ollama = PIG_PATH / "llm_answer" / "answer_slicing.xlsx"
    else:
        ollama = PIG_PATH / "llm_answer" / "answer_pig.xlsx"
    

    if model == "llama3.1-8b":
        load_wb = load_workbook(ollama, data_only=True)
        load_ws = load_wb["llama3.1-8b"]

    elif model == "gemma2-9b":
        load_wb = load_workbook(ollama, data_only=True)
        load_ws = load_wb["gemma2-9b"]

    elif model == "qwen2-7b":
        load_wb = load_workbook(ollama, data_only=True)
        load_ws = load_wb["qwen2-7b"]

    elif model == "qwen3-32b":
        load_wb = load_workbook(ollama, data_only=True)
        load_ws = load_wb["qwen3-32b"]

    elif model == "deepseek-r1-32b":
        load_wb = load_workbook(ollama, data_only=True)
        load_ws = load_wb["deepseek-r1-32b"]

    elif model == "gemma3-27b":
        load_wb = load_workbook(ollama, data_only=True)
        load_ws = load_wb["gemma3-27b"]

    else:
        print("Invalid model name")
        return None

    result = dict()

    for row in load_ws.rows:
        _id = row[0].value
        result[_id] = [
            row[i].value.strip("\n")
            for i in range(2, len(row), 2)
            if row[i].value != None
        ]

    return result


def CodeExtract(llm_answer: str, libo, libn) -> list[str]:
    # Extract code snippet
    result = list()
    start = None

    for i in range(len(llm_answer)):
        if (start == None) and (
            llm_answer[i] == "`" and llm_answer[i : i + 9] == "```python"
        ):
            start = i + 9

        elif (start == None) and (
            llm_answer[i] == "`" and llm_answer[i : i + 3] == "```"
        ):
            start = i + 2

        elif (start != None) and llm_answer[i : i + 3] == "```":
            end = i
            result.append(llm_answer[start:end].strip("`"))
            start = None

    if "```" not in llm_answer:
        result.append(llm_answer)

    if len(result) == 1:
        try:
            ast.unparse(libn)

        except:
            if libn in result[0]:
                result[0] = result[0].replace(libn, llm_pre.libname(libn))

        return result

    # Unchanged
    for i in range(len(result)):
        answer: str = result[i]
        answers = ""

        for j in range(len(answer.split("\n"))):
            line = answer.split("\n")[j]

            if "unchanged" in line and "#" in line:
                answers += line.replace("#", "pass #") + "\n"

            elif "existing" in line and "#" in line:
                answers += line.replace("#", "pass #") + "\n"

            elif "code continues" in line and "#" in line:
                answers += line.replace("#", "pass #") + "\n"

            elif "Do something" in line and "#" in line:
                answers += line.replace("#", "pass #") + "\n"

            elif "..." in line:
                answers += line.replace("...", "dummy_var") + "\n"

            else:
                answers += line + "\n"

        result[i] = answers

    if len(result) == 0 and "```python" in llm_answer:
        result.append(llm_answer[start:].strip())

    # if libo exists~
    for i in range(len(result)):
        answer = result[i]

        for j in range(len(answer.split("\n"))):
            line = answer.split("\n")[j]

            if "import" in line:
                try:
                    imp_line: Union[ast.Import, ast.ImportFrom] = ast.parse(
                        line.strip()
                    ).body[0]
                except:
                    continue

                libraries = set()

                if isinstance(imp_line, ast.Import):
                    for alias in imp_line.names:
                        libraries.add(alias.name.split(".")[0])

                elif isinstance(imp_line, ast.ImportFrom) and imp_line.module != None:
                    libraries.add(imp_line.module.split(".")[0])

                # ruamel.yaml....
                if llm_pre.libname(libo) in libraries and (
                    not llm_pre.libname(libn) in line
                ):
                    result[i] = ""
                    break

    return result